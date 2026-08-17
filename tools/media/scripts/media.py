"""Built-in media family — audio, video, documents (one tool).

The Operator's 08-12 spec: a `media` tool that handles audio, video, and
documents. Uses ffprobe (available on the host) for media files and
plain file inspection for documents. All read-only — reports metadata,
never transcode (that stays out of scope for a hands-off button).
"""

import json
import mimetypes
import subprocess
from pathlib import Path


def _ffprobe(path: str) -> dict | None:
    """Run ffprobe on a media file, return the JSON stream info."""
    try:
        r = subprocess.run(
            ["ffprobe", "-v", "quiet", "-print_format", "json",
             "-show_format", "-show_streams", path],
            capture_output=True, text=True, timeout=30)
        if r.returncode != 0:
            return None
        return json.loads(r.stdout)
    except Exception:
        return None


def _probe(args: dict, timeout: float = 15.0) -> str:
    path = str(args.get("path", "")).strip()
    if not path:
        return json.dumps({"ok": False, "detail": "path required"},
                          ensure_ascii=False)
    p = Path(path).expanduser()
    if not p.exists():
        return json.dumps({"ok": False, "detail": f"not found: {path}"},
                          ensure_ascii=False)
    if not p.is_file():
        return json.dumps({"ok": False, "detail": f"not a file: {path}"},
                          ensure_ascii=False)
    data = _ffprobe(str(p))
    if data:
        fmt = data.get("format", {})
        streams = data.get("streams", [])
        video = next((s for s in streams if s.get("codec_type") == "video"), None)
        audio = next((s for s in streams if s.get("codec_type") == "audio"), None)
        return json.dumps({
            "ok": True,
            "path": str(p),
            "size_bytes": p.stat().st_size,
            "format": fmt.get("format_long_name") or fmt.get("format_name"),
            "duration_s": round(float(fmt.get("duration", 0)), 2),
            "bitrate_kbps": round(int(fmt.get("bit_rate", 0)) / 1000) if fmt.get("bit_rate") else 0,
            "video": {
                "codec": (video or {}).get("codec_name"),
                "width": (video or {}).get("width"),
                "height": (video or {}).get("height"),
                "fps": (video or {}).get("avg_frame_rate"),
            } if video else None,
            "audio": {
                "codec": (audio or {}).get("codec_name"),
                "channels": (audio or {}).get("channels"),
                "sample_rate": (audio or {}).get("sample_rate"),
            } if audio else None,
        }, ensure_ascii=False)
    # Not a media file — document/fallback inspection.
    mime, _ = mimetypes.guess_type(path)
    try:
        head = p.read_bytes()[:16]
        kind = "unknown"
        if head[:4] == b"%PDF":
            kind = "pdf"
        elif head[:8] == b"PK\x03\x04":
            kind = "zip/office (docx/xlsx/pptx)"
        elif head[:4] == b"\x89PNG":
            kind = "png"
        elif head[:2] == b"\xff\xd8":
            kind = "jpeg"
        elif head[:4] == b"GIF8":
            kind = "gif"
        elif head[:4] == b"OggS":
            kind = "ogg"
        elif head[:4] == b"RIFF":
            kind = "avi/wav"
        return json.dumps({
            "ok": True,
            "path": str(p),
            "size_bytes": p.stat().st_size,
            "mime": mime or "application/octet-stream",
            "kind": kind,
        }, ensure_ascii=False)
    except Exception as exc:
        return json.dumps({"ok": False, "detail": str(exc)}, ensure_ascii=False)


def _image(args: dict, timeout: float = 15.0) -> str:
    """Rich image info via Pillow: dimensions, mode, format, EXIF."""
    path = str(args.get("path", "")).strip()
    if not path:
        return json.dumps({"ok": False, "detail": "path required"},
                          ensure_ascii=False)
    p = Path(path).expanduser()
    if not p.is_file():
        return json.dumps({"ok": False, "detail": f"not found: {path}"},
                          ensure_ascii=False)
    try:
        from PIL import Image, ExifTags
        with Image.open(str(p)) as im:
            exif = {}
            try:
                for k, v in im.getexif().items():
                    name = ExifTags.TAGS.get(k, str(k))
                    exif[name] = str(v)[:80]
            except Exception:
                pass
            return json.dumps({
                "ok": True,
                "path": str(p),
                "size_bytes": p.stat().st_size,
                "format": im.format,
                "mode": im.mode,
                "width": im.width,
                "height": im.height,
                "animated": getattr(im, "is_animated", False),
                "frames": getattr(im, "n_frames", 1),
                "exif": exif or None,
            }, ensure_ascii=False)
    except Exception as exc:
        return json.dumps({"ok": False, "detail": f"image error: {exc}"},
                          ensure_ascii=False)


def _pdf(args: dict, timeout: float = 30.0) -> str:
    """PDF operations: pages, extract text, merge, split (pypdf)."""
    path = str(args.get("path", "")).strip()
    op = str(args.get("pdf_action", "info")).strip()
    p = Path(path).expanduser() if path else None
    if p and not p.is_file():
        return json.dumps({"ok": False, "detail": f"not found: {path}"},
                          ensure_ascii=False)
    try:
        from pypdf import PdfReader, PdfWriter
        if op == "info":
            reader = PdfReader(str(p))
            return json.dumps({
                "ok": True, "path": str(p), "kind": "pdf",
                "pages": len(reader.pages),
                "size_bytes": p.stat().st_size,
                "metadata": {k: str(v)[:100] for k, v in
                             (reader.metadata or {}).items()},
            }, ensure_ascii=False)
        if op == "text":
            reader = PdfReader(str(p))
            page = int(args.get("page", 1)) - 1
            if 0 <= page < len(reader.pages):
                text = reader.pages[page].extract_text() or ""
                return json.dumps({"ok": True, "path": str(p),
                                   "page": page + 1, "text": text[:3000]},
                                  ensure_ascii=False)
            return json.dumps({"ok": False, "detail": "page out of range"},
                              ensure_ascii=False)
        if op == "merge":
            out = str(args.get("out", "")).strip()
            paths = args.get("paths") or []
            if not out or not paths:
                return json.dumps({"ok": False,
                                   "detail": "out and paths required"},
                                  ensure_ascii=False)
            writer = PdfWriter()
            for f in paths:
                writer.append(str(Path(str(f)).expanduser()))
            with open(out, "wb") as fh:
                writer.write(fh)
            return json.dumps({"ok": True, "created": out,
                               "sources": len(paths)}, ensure_ascii=False)
        if op == "split":
            out_dir = str(args.get("out", "")).strip()
            if not out_dir:
                return json.dumps({"ok": False, "detail": "out dir required"},
                                  ensure_ascii=False)
            od = Path(out_dir).expanduser()
            od.mkdir(parents=True, exist_ok=True)
            reader = PdfReader(str(p))
            made = []
            for i, page in enumerate(reader.pages, 1):
                writer = PdfWriter()
                writer.add_page(page)
                f = od / f"page-{i}.pdf"
                with open(f, "wb") as fh:
                    writer.write(fh)
                made.append(f.name)
            return json.dumps({"ok": True, "pages": len(made), "files": made},
                              ensure_ascii=False)
        return json.dumps({"ok": False, "detail": f"unknown pdf op: {op}"},
                          ensure_ascii=False)
    except Exception as exc:
        return json.dumps({"ok": False, "detail": f"pdf error: {exc}"},
                          ensure_ascii=False)


def _table(args: dict, timeout: float = 30.0) -> str:
    """Row/column documents: xlsx, csv, sqlite (the Operator's spec)."""
    path = str(args.get("path", "")).strip()
    if not path:
        return json.dumps({"ok": False, "detail": "path required"},
                          ensure_ascii=False)
    p = Path(path).expanduser()
    if not p.is_file():
        return json.dumps({"ok": False, "detail": f"not found: {path}"},
                          ensure_ascii=False)
    limit = int(args.get("limit", 20))
    suffix = p.suffix.lower()
    try:
        if suffix == ".csv":
            import csv
            with open(p, newline="", encoding="utf-8", errors="replace") as f:
                rows = list(csv.reader(f))
            header = rows[0] if rows else []
            return json.dumps({
                "ok": True, "kind": "csv", "path": str(p),
                "columns": header, "rows": len(rows) - 1,
                "sample": rows[: min(limit + 1, len(rows))],
            }, ensure_ascii=False)
        if suffix in (".xlsx", ".xlsm"):
            import openpyxl
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            ws = wb.active
            header = [c.value for c in next(ws.iter_rows(max_row=1))]
            total = ws.max_row - 1
            sample = [[c.value for c in row]
                      for row in ws.iter_rows(min_row=2,
                                              max_row=min(limit + 1, ws.max_row))]
            wb.close()
            return json.dumps({
                "ok": True, "kind": "xlsx", "path": str(p),
                "sheet": ws.title, "columns": header,
                "rows": max(total, 0), "sample": sample,
            }, ensure_ascii=False)
        if suffix in (".sqlite", ".sqlite3", ".db"):
            import sqlite3
            conn = sqlite3.connect(str(p))
            tables = [r[0] for r in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'")]
            table = str(args.get("table", "")).strip() or (tables[0] if tables else "")
            if not table:
                conn.close()
                return json.dumps({"ok": True, "kind": "sqlite", "path": str(p),
                                   "tables": tables}, ensure_ascii=False)
            cols = [d[1] for d in conn.execute(f"PRAGMA table_info({table})")]
            rows = conn.execute(f"SELECT * FROM {table} LIMIT {limit}").fetchall()
            total = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            conn.close()
            return json.dumps({
                "ok": True, "kind": "sqlite", "path": str(p),
                "table": table, "tables": tables, "columns": cols,
                "rows": total, "sample": rows,
            }, ensure_ascii=False)
        return json.dumps({"ok": False,
                           "detail": f"unsupported table format: {suffix}"},
                          ensure_ascii=False)
    except Exception as exc:
        return json.dumps({"ok": False, "detail": f"table error: {exc}"},
                          ensure_ascii=False)


def _media(args: dict, timeout: float = 15.0) -> str:
    action = str(args.get("action", "")).strip()
    if action in ("probe", "info", "duration", "document"):
        return _probe(args, timeout)
    if action == "image":
        return _image(args, timeout)
    if action == "pdf":
        return _pdf(args, timeout)
    if action == "table":
        return _table(args, timeout)
    return json.dumps({"ok": False, "detail": f"unknown action: {action}"},
                      ensure_ascii=False)


def register() -> list[str]:
    from filesystem.tools import Tool, register
    register(Tool(
        name="media",
        description="Media tool (the Operator's 08-12 spec): probe audio/video/"
                    "documents, rich image info (Pillow), PDF operations "
                    "(pages/text/merge/split via pypdf), and table docs "
                    "(xlsx/csv/sqlite rows+columns). Read-only where "
                    "probing; PDF/table ops write to explicit out paths.",
        parameters={
            "type": "object",
            "properties": {
                "action": {"type": "string",
                           "enum": ["probe", "info", "duration", "document",
                                    "image", "pdf", "table"]},
                "path": {"type": "string", "description": "File path"},
                "pdf_action": {"type": "string",
                               "enum": ["info", "text", "merge", "split"]},
                "page": {"type": "integer", "description": "PDF page (1-based)"},
                "paths": {"type": "array", "items": {"type": "string"},
                          "description": "PDFs to merge"},
                "out": {"type": "string", "description": "Output path/dir"},
                "table": {"type": "string", "description": "SQLite table name"},
                "limit": {"type": "integer", "description": "Row limit"},
            },
            "required": ["action"],
        },
        fn=_media,
    ))
    return ["media"]
